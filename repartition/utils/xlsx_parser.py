import openpyxl

# Mapping semestre → plages de lignes dans l'onglet Maquette (1-indexé)
SEMESTRES_MAQUETTE = {
    "S1":      (3, 27),
    "S2":      (39, 64),
    "S3":      (76, 98),
    "S4 crea": (110, 123),
    "S4 dev":  (136, 148),
    "S5 crea": (160, 170),
    "S5 dev":  (182, 195),
    "S6 crea": (209, 213),
    "S6 dev":  (226, 231),
}

# Mots-clés à ignorer dans la colonne D
IGNORE_KEYWORDS = ["total", "contrainte", "intitulé", "heures ptut", "adaptation"]


def _is_valid_ressource(value):
    if not value:
        return False
    v = str(value).strip().lower()
    if not v:
        return False
    for kw in IGNORE_KEYWORDS:
        if v.startswith(kw):
            return False
    return True


def get_ressources_semestre(xlsx_path, semestre):
    """
    Retourne la liste des intitulés de ressources/SAÉ/Portfolio
    pour un semestre donné, lus depuis l'onglet Maquette colonne D.
    """
    if semestre not in SEMESTRES_MAQUETTE:
        raise ValueError(f"Semestre inconnu : {semestre}")

    row_min, row_max = SEMESTRES_MAQUETTE[semestre]
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["Maquette"]

    ressources = []
    for row in ws.iter_rows(min_row=row_min, max_row=row_max, values_only=True):
        val = row[3] if len(row) > 3 else None  # colonne D = index 3
        if _is_valid_ressource(val):
            ressources.append(str(val).strip())

    wb.close()
    return ressources


def get_enseignants(xlsx_path):
    """
    Retourne la liste des enseignants depuis l'onglet ListeEnseignants.
    Format : [{"id": "NOM Prenom", "nom": "NOM", "prenom": "Prenom",
               "service_du": 384, "service_max": None, "is_vac": False}, ...]
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["ListeEnseignants"]

    enseignants = []
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i == 1:
            continue  # ligne d'en-têtes
        nom = row[0] if len(row) > 0 else None
        prenom = row[1] if len(row) > 1 else None
        if not nom and not prenom:
            continue
        nom = str(nom).strip() if nom else ""
        prenom = str(prenom).strip() if prenom else ""
        if not nom and not prenom:
            continue

        # service_du peut être une formule string genre '=0.8*384'
        raw_service_du = row[3] if len(row) > 3 else None
        service_du = None
        if raw_service_du is not None:
            try:
                service_du = float(raw_service_du)
            except (ValueError, TypeError):
                try:
                    # évaluation sécurisée de formules simples type =0.8*384
                    expr = str(raw_service_du).lstrip("=")
                    service_du = float(eval(expr))
                except Exception:
                    service_du = None

        raw_service_max = row[4] if len(row) > 4 else None
        service_max = None
        if raw_service_max is not None:
            try:
                service_max = float(raw_service_max)
            except (ValueError, TypeError):
                try:
                    expr = str(raw_service_max).lstrip("=")
                    service_max = float(eval(expr))
                except Exception:
                    service_max = None

        raw_is_vac = row[5] if len(row) > 5 else "False"
        is_vac = str(raw_is_vac).strip().lower() == "true"

        uid = f"{nom.upper()} {prenom}"

        enseignants.append({
            "id": uid,
            "nom": nom.upper(),
            "prenom": prenom,
            "service_du": service_du,
            "service_max": service_max,
            "is_vac": is_vac,
        })

    wb.close()
    enseignants.sort(key=lambda e: (e["nom"], e["prenom"]))
    return enseignants


def get_all_ressources(xlsx_path):
    """Retourne un dict {semestre: [ressources]} pour tous les semestres."""
    result = {}
    for semestre in SEMESTRES_MAQUETTE:
        result[semestre] = get_ressources_semestre(xlsx_path, semestre)
    return result


def get_maquette_data(xlsx_path, semestre):
    """
    Retourne les données colonnes D à K pour un semestre depuis l'onglet Maquette.
    Colonnes (0-indexées) : D=3, E=4, F=5, G=6, H=7, I=8, J=9, K=10
    Retourne une liste de dicts par ressource valide.
    """
    if semestre not in SEMESTRES_MAQUETTE:
        raise ValueError(f"Semestre inconnu : {semestre}")

    row_min, row_max = SEMESTRES_MAQUETTE[semestre]
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["Maquette"]

    def _num(v):
        if v is None:
            return 0
        try:
            return float(v)
        except (ValueError, TypeError):
            return 0

    rows = []
    for row in ws.iter_rows(min_row=row_min, max_row=row_max, values_only=True):
        intitule = row[3] if len(row) > 3 else None
        if not _is_valid_ressource(intitule):
            continue
        rows.append({
            "intitule":     str(intitule).strip(),
            "vol_national": _num(row[4] if len(row) > 4 else None),
            "dont_tp_nat":  _num(row[5] if len(row) > 5 else None),
            "adapt_locale": _num(row[6] if len(row) > 6 else None),
            "dont_tp_adapt":_num(row[7] if len(row) > 7 else None),
            "cm_final":     _num(row[8] if len(row) > 8 else None),
            "td_final":     _num(row[9] if len(row) > 9 else None),
            "tp_final":     _num(row[10] if len(row) > 10 else None),
        })

    wb.close()
    return rows
